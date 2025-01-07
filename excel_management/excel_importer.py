""" © Daniel P Raven and Matt Russell 2024 All Rights Reserved """
import openpyxl
from qtr_pairing_process.excel_management.constants import SHEET_NAMES
from qtr_pairing_process.db_management.db_manager import DbManager
from qtr_pairing_process.constants import SCENARIO_MAP
class ExcelImporter:
    def __init__(self, db_manager: DbManager, file_path, file_name):
        self.workbook = None
        self.team_metadata = {}
        self.player_dicts = {}
        self.db_manager = db_manager
        self.ratings = {}
        self.file_path = file_path
        self.file_name = file_name

    def read_excel_file(self):
        print(self.file_path, self.file_name)
        self.workbook = openpyxl.load_workbook(filename = f'{self.file_path}/{self.file_name}')

    def validate_sheetnames(self):
        workbook_sheet_names = self.workbook.sheetnames
        for sn in SHEET_NAMES:
            if sn not in workbook_sheet_names:
                print(f'Sheet {sn} not found')

    def validate_and_assign_team(self, team_position, team_name, player_names):
        if not isinstance(team_name, str) or team_name=='':
            print(f'Invalid team name for {team_position}: {team_name}')
        for player_name in player_names:
            if not isinstance(player_name, str) or player_name=='':
                print(f'Invalid player_name for {team_position}: {player_name}')

        team_dict = {
            'team_name': team_name,
            'player_names': player_names
        }
        self.team_metadata[team_position] = team_dict

    def validate_team_sheet(self, team_sheet):
        team1 = [x[0].value for x in team_sheet['A1':'A6']]
        team1_name = team1[0]
        team1_players = team1[1:]
        self.validate_and_assign_team('team1',team1_name, team1_players)

        team2 = [x[0].value for x in team_sheet['B1':'B6']]
        team2_name = team2[0]
        team2_players = team2[1:]
        self.validate_and_assign_team('team2',team2_name, team2_players)


    def validate_ranking_sheet_players(self, sheet_name):
        sheet = self.workbook[sheet_name]
        team1_ranking_players = [x[0].value for x in sheet['A2':'A6']]
        team2_ranking_players = [x.value for x in sheet['B1':'F1'][0]]

        if team1_ranking_players != self.team_metadata['team1']['player_names']:
            print(f'Varying team row names between team sheet and sheet_name {sheet_name}.')

        if team2_ranking_players != self.team_metadata['team2']['player_names']:
            print(f'Varying team column names between team sheet and sheet_name {sheet_name}.')
    def validate_ranking_sheet_ranks(self, sheet_name):
        sheet = self.workbook[sheet_name]
        grid_ranks = sheet['B2':'F6']
        for row in grid_ranks:
            for col in row:
                if not isinstance(col.value,int):
                    if isinstance(col.value,float):
                        if col.value != int(col.value):
                            print(f'Floats not valid for ranking, should be INT in range 1-5: {col.value}')
                    else:    
                        print(f'Not a valid value for ranking: {col.value}')
                elif col.value < 1 or col.value > 5:
                    print(f'Bad value {col.value} for rank, should be in range 1-5.')

    def read_ranking_sheet(self, sheet_name):
        sheet = self.workbook[sheet_name]
        grid_ranks = sheet['B2':'F6']
        grid_rank_array = [[x.value for x in y] for y in grid_ranks]
        self.ratings[sheet_name] = grid_rank_array

    def read_and_validate_ranking_sheets(self):
        ranking_sheet_names = SHEET_NAMES[1:]
        for sheet_name in ranking_sheet_names:
            self.validate_ranking_sheet_players(sheet_name)
            self.validate_ranking_sheet_ranks(sheet_name)
        
        for sheet_name in ranking_sheet_names:
            self.read_ranking_sheet(sheet_name)

    def upsert_rating_sheet(self, rating_sheet_name, scenario_id):
        grid_ranks = self.ratings[rating_sheet_name]
        team_1 = self.team_metadata['team1']
        team_2 = self.team_metadata['team2']
        team_1_id = self.db_manager.query_team_id(team_1['team_name'])
        team_2_id = self.db_manager.query_team_id(team_2['team_name'])

        if team_1_id > team_2_id:
            team_1_id, team_2_id = team_2_id, team_1_id
            team_1, team_2 = team_2, team_1



        for i,row in enumerate(grid_ranks):
            for j,col in enumerate(row):
                
                player_id_1 = self.player_dicts['team1'][self.team_metadata['team1']['player_names'][i]]
                player_id_2 = self.player_dicts['team2'][self.team_metadata['team2']['player_names'][j]]
                rank = col
                self.db_manager.upsert_rating(player_id_1=player_id_1,
                player_id_2=player_id_2,
                team_id_1=team_1_id,
                team_id_2=team_2_id,
                scenario_id=scenario_id,
                rating=rank,
                )

    def import_team_and_player_data_to_db(self, team_name, player_names, team_position):
        team_id = self.db_manager.upsert_team(team_name=team_name)
        players = self.db_manager.upsert_and_validate_players(team_id=team_id, player_names=player_names)
        player_dict = {x[1]:x[0] for x in players}
        self.player_dicts[team_position] = player_dict

    def execute(self):
        print('Beginning Excel Import!!!')
        self.read_excel_file()
        self.validate_sheetnames()

        self.validate_team_sheet(self.workbook[SHEET_NAMES[0]])
        self.read_and_validate_ranking_sheets()

        self.import_team_and_player_data_to_db(
            team_name=self.team_metadata['team1']['team_name'],
            player_names=self.team_metadata['team1']['player_names'],
            team_position='team1',)
        
        self.import_team_and_player_data_to_db(
            team_name=self.team_metadata['team2']['team_name'],
            player_names=self.team_metadata['team2']['player_names'],
            team_position='team2',)
        for scenario_id, scenario_name in SCENARIO_MAP.items():
            self.upsert_rating_sheet(rating_sheet_name=scenario_name, scenario_id=scenario_id)
